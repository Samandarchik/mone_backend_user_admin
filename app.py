#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Print Server with Error Handling & Telegram Notifications
Version: 2.3 - Added Error Logging + Telegram Bot Integration
"""

from flask import Flask, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl import Workbook
import os
from datetime import datetime
import tempfile
import traceback
import requests

app = Flask(__name__)

# Telegram bot settings
TELEGRAM_BOT_TOKEN = "8157743798:AAELzxyyFLSMxbT-XL4l-3ZVmxVBXYOY0Ro"
TELEGRAM_CHAT_ID = "1066137436"

# Printer settings
PRINTERS = {
    1: "Canon LBP6030",
    2: "HP LaserJet 1020", 
    3: "Epson L3150",
    4: "Brother HL-1110"
}

def send_telegram_error(error_message, endpoint="", request_data=None):
    """Send error notification to Telegram"""
    try:
        message = f"🚨 <b>Excel Print Server Error</b>\n\n"
        message += f"⏰ Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        message += f"📍 Endpoint: {endpoint}\n"
        message += f"❌ Error: {error_message}\n"
        
        if request_data:
            message += f"\n📦 Request Data:\n<code>{str(request_data)[:500]}</code>\n"
        
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {
            "chat_id": TELEGRAM_CHAT_ID,
            "text": message,
            "parse_mode": "HTML"
        }
        
        response = requests.post(url, json=payload, timeout=5)
        if response.status_code == 200:
            print("✅ Error sent to Telegram")
        else:
            print(f"⚠️ Failed to send to Telegram: {response.status_code}")
            
    except Exception as e:
        print(f"❌ Telegram notification failed: {e}")

def validate_items(items):
    """Validate items data structure"""
    if not items or not isinstance(items, list):
        raise ValueError("Items must be a non-empty list")
    
    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            raise ValueError(f"Item {idx} must be a dictionary")
        
        # Check required fields
        if 'name' not in item and 'product' not in item:
            raise ValueError(f"Item {idx} missing 'name' or 'product' field")
        
        # Validate count
        count = item.get('count', 0)
        try:
            float(count) if count is not None else 0
        except (ValueError, TypeError):
            raise ValueError(f"Item {idx} has invalid 'count' value: {count}")
    
    return True

def create_excel_file(items, username=None, filial=None, order_id=None, category=None):
    """Create Excel file with error handling"""
    try:
        # Validate items first
        validate_items(items)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказ"
        
        # Page setup for A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Page margins (in inches)
        ws.page_margins = PageMargins(
            left=0.2, right=0.2, top=0.2, bottom=0.2,
            header=0.3, footer=0.3
        )
        
        # Styles
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=11)
        total_font = Font(name='Arial', size=12, bold=True)
        info_font = Font(name='Arial', size=12, bold=True)
        
        # Colors
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = 'СПИСОК ЗАКАЗОВ: ' + (order_id or 'N/A')
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        row = 3

        # Category
        if category:
            ws[f'A{row}'] = 'Категория:  ' + str(category)
            ws[f'A{row}'].font = info_font
            ws.merge_cells(f'A{row}:C{row}')
            row += 1

        # Customer
        if username:
            ws[f'A{row}'] = 'Заказчик:  ' + str(username)
            ws[f'A{row}'].font = info_font
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        # Branch
        if filial:
            ws[f'A{row}'] = 'Филиал:  ' + str(filial)
            ws[f'A{row}'].font = info_font
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        # Time
        ws[f'A{row}'] = 'Время:   ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = info_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 2
        
        # Table header
        ws[f'A{row}'] = '№'
        ws[f'B{row}'] = 'Название товара'
        ws[f'C{row}'] = 'Ед. изм.'
        ws[f'D{row}'] = 'Кол-во'
        
        # Apply header style
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row += 1
        start_data_row = row
        
        # Items
        total_count = 0
        for i, item in enumerate(items, 1):
            product = item.get('name', item.get('product', 'Неизвестно'))
            count = item.get('count', 0)
            type_unit = item.get('type', 'шт')
            
            # Handle double/float values
            try:
                count_value = float(count) if count is not None else 0
            except (ValueError, TypeError):
                count_value = 0
                print(f"⚠️ Warning: Invalid count for item {i}: {count}")
            
            total_count += count_value
            
            ws[f'A{row}'] = i
            ws[f'B{row}'] = str(product)
            ws[f'C{row}'] = str(type_unit)
            ws[f'D{row}'] = count_value
            
            # Number format for double values
            ws[f'D{row}'].number_format = '0.00'
            
            # Styles for data rows
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Product name - large font
            ws[f'B{row}'].font = Font(name='Arial', size=13.5, bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Unit of measurement
            ws[f'C{row}'].font = normal_font
            ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Quantity
            ws[f'D{row}'].font = normal_font
            ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Borders for all cells
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = thin_border

            # Row height
            ws.row_dimensions[row].height = 25
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = ''
        ws[f'B{row}'] = ''
        ws[f'C{row}'] = 'ИТОГО:'
        ws[f'D{row}'] = total_count
        
        ws[f'D{row}'].number_format = '0.00'
        
        # Apply total row style
        ws[f'C{row}'].font = total_font
        ws[f'D{row}'].font = total_font
        ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].border = thin_border
        
        ws.row_dimensions[row].height = 29
        
        # Column widths for A4 format
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 65
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        
        # Print settings
        ws.print_options.horizontalCentered = False
        ws.print_options.verticalCentered = False
        
        # Repeat headers on each page
        ws.print_title_rows = f'1:{start_data_row-1}'
        
        # Filename and path
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"zakaz_{timestamp}.xlsx"
        
        # Save to temp folder
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        
        # Save
        wb.save(filepath)
        
        print(f"✅ Excel file created: {filename}")
        return filepath, filename
        
    except Exception as e:
        error_msg = f"Error creating Excel: {str(e)}\n{traceback.format_exc()}"
        print(f"❌ {error_msg}")
        send_telegram_error(error_msg, "create_excel_file")
        raise

def print_excel_file(filepath, printer_name):
    """Print Excel file"""
    try:
        import subprocess
        import sys
        
        if sys.platform == "win32":
            os.startfile(filepath, "print")
            print(f"✅ Excel sent to printer: {printer_name}")
            return True
        else:
            print("❌ Only works on Windows")
            return False
            
    except Exception as e:
        error_msg = f"Print error: {str(e)}"
        print(f"❌ {error_msg}")
        send_telegram_error(error_msg, "print_excel_file")
        return False

@app.route('/', methods=['GET'])
def api_info():
    """API Information"""
    return jsonify({
        "name": "Excel Print Server",
        "version": "2.3",
        "status": "Working",
        "description": "Creates and prints Excel files with error notifications",
        "printers": PRINTERS,
        "features": [
            "Excel (.xlsx) file creation",
            "Professional formatting",
            "A4 print optimization",
            "Category support",
            "Unit of measurement (type)",
            "Printing",
            "Download",
            "Double number support",
            "Telegram error notifications",
            "Data validation"
        ],
        "endpoints": {
            "POST /print": "Create and print Excel",
            "POST /excel": "Create Excel only (for download)",
            "GET /": "API information",
            "GET /test": "Create test Excel"
        }
    })

@app.route('/print', methods=['POST'])
def api_print():
    """Create and print Excel"""
    try:
        data = request.get_json()
        if not data:
            error = "JSON data required"
            send_telegram_error(error, "/print", data)
            return jsonify({"error": error}), 400
        
        printer_key = data.get('printer', 1)
        items = data.get('items', [])
        username = data.get('username')
        filial = data.get('filial')
        order_id = data.get('order_id')
        category = data.get('category')
        
        if not items:
            error = "Items list is empty"
            send_telegram_error(error, "/print", data)
            return jsonify({"error": error}), 400
        
        if printer_key not in PRINTERS:
            error = f"Printer not found: {list(PRINTERS.keys())}"
            send_telegram_error(error, "/print", data)
            return jsonify({"error": error}), 400
        
        printer_name = PRINTERS[printer_key]
        
        # Create Excel file
        filepath, filename = create_excel_file(items, username, filial, order_id, category)
        if not filepath:
            error = "Failed to create Excel"
            send_telegram_error(error, "/print", data)
            return jsonify({"error": error}), 500
        
        # Print
        # print_success = print_excel_file(filepath, printer_name)
        
        response = {
            "success": True,
            "message": f"Excel created and sent to {printer_name}",
            "filename": filename,
            "printer": printer_name,
            "items_count": len(items),
            "total_quantity": sum(float(item.get('count', 0)) if item.get('count') is not None else 0 for item in items),
            "print_status": "sent" if True else "print error",
            "file_path": filepath,
            "order_id": order_id,
            "category": category
        }
        
        return jsonify(response)
        
    except Exception as e:
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        send_telegram_error(error_msg, "/print", request.get_json())
        return jsonify({"error": str(e)}), 500

@app.route('/excel', methods=['POST'])
def api_excel_only():
    """Create Excel only (for download)"""
    try:
        data = request.get_json()
        if not data:
            error = "JSON data required"
            send_telegram_error(error, "/excel", data)
            return jsonify({"error": error}), 400
        
        items = data.get('items', [])
        username = data.get('username')
        filial = data.get('filial')
        order_id = data.get('order_id')
        category = data.get('category')
        
        if not items:
            error = "Items list is empty"
            send_telegram_error(error, "/excel", data)
            return jsonify({"error": error}), 400
        
        # Create Excel file
        filepath, filename = create_excel_file(items, username, filial, order_id, category)
        if not filepath:
            error = "Failed to create Excel"
            send_telegram_error(error, "/excel", data)
            return jsonify({"error": error}), 500
        
        # Send file for download
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        send_telegram_error(error_msg, "/excel", request.get_json())
        return jsonify({"error": str(e)}), 500

@app.route('/test', methods=['GET'])
def api_test():
    """Create test Excel"""
    test_items = [
        {"product_id": 90, "name": "Кортошка оллади", "count": 4.5, "type": "шт"},
        {"product_id": 91, "name": "Молоко свежее", "count": 2.75, "type": "л"},
        {"product_id": 92, "name": "Хлеб белый", "count": 3.25, "type": "шт"},
        {"product_id": 93, "name": "Сахар", "count": 1.5, "type": "кг"},
        {"product_id": 94, "name": "Масло подсолнечное", "count": 1.2, "type": "л"}
    ]
    
    try:
        filepath, filename = create_excel_file(
            test_items, 
            "Тестовый пользователь", 
            "Тестовый филиал", 
            "TEST-001",
            "Продукты питания"
        )
        
        if filepath:
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({"error": "Failed to create test Excel"}), 500
    except Exception as e:
        send_telegram_error(str(e), "/test")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    print("=== EXCEL PRINT SERVER (WITH ERROR NOTIFICATIONS) ===")
    print("URL: http://localhost:2020")
    print("Test Excel: http://localhost:2020/test")
    print(f"Telegram notifications: {TELEGRAM_CHAT_ID}")
    print("Printers:", list(PRINTERS.keys()))
    print("Ctrl+C - Stop")
    print("===================================================")
    
    # Check openpyxl library
    try:
        import openpyxl
        print("✅ openpyxl library available")
    except ImportError:
        print("❌ openpyxl library not found!")
        print("💡 Install: pip install openpyxl")
    
    # Check requests library
    try:
        import requests
        print("✅ requests library available")
    except ImportError:
        print("❌ requests library not found!")
        print("💡 Install: pip install requests")
    
    app.run(host='0.0.0.0', port=2020, debug=False)